import sys
import openpyxl
from bisect import bisect_right
from datetime import datetime, date
from PyQt6.QtWidgets import QApplication, QMessageBox


SHEET_NAME = "overwrite_registry"
FILTER_Q_VALUES_1 = {
    "Overwrite I - Sobrescritura de FlamaNova & HormaNova",
    "Overwrite I - Handler de Dorothy | Post FlamaNova",
}
FILTER_Q_VALUES_2 = {
    "Overwrite II - Le Etude de Dorothée",
    "Overwrite II - Sobrescritura de Dorothy & Lissette",
}
FILTER_E_VALUE = "Episodio"


def msgbox(title: str, message: str, type_icon=QMessageBox.Icon.Information) -> None:
    msg = QMessageBox()
    msg.setIcon(type_icon)
    msg.setWindowTitle(title)
    msg.setText(message)
    msg.exec()


def _to_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        raw = value.strip()
        if len(raw) >= 10:
            try:
                return datetime.strptime(raw[:10], "%Y-%m-%d").date()
            except ValueError:
                return None
    return None


def _xlookup_last_le(target: date, rows: list[tuple[date, object]]) -> object:
    """XLOOKUP(..., match_mode=-1) equivalent for sorted date keys."""
    keys = [k for k, _ in rows]
    idx = bisect_right(keys, target) - 1
    if idx < 0:
        return "-"
    return rows[idx][1]


def _xlookup_last_le_pair(target: date, rows: list[tuple[date, object, object]]) -> tuple[object, object]:
    """Return (principal, name) for the last date <= target."""
    keys = [k for k, _, _ in rows]
    idx = bisect_right(keys, target) - 1
    if idx < 0:
        return "-", "-"
    return rows[idx][1], rows[idx][2]


def _build_material_map(ws) -> dict[str, object]:
    """
    material_list: title in column I, type in column E
    """
    mapping: dict[str, object] = {}
    for r in range(4, ws.max_row + 1):
        title = ws[f"I{r}"].value
        typ = ws[f"E{r}"].value
        if title is None:
            continue
        title = str(title).strip()
        if title and title not in mapping:
            mapping[title] = typ
    return mapping


def _build_ov_lookup(ws, start_col: str, principal_col: str, name_col: str) -> list[tuple[date, object, object]]:
    rows: list[tuple[date, object, object]] = []
    for r in range(3, ws.max_row + 1):
        start = _to_date(ws[f"{start_col}{r}"].value)
        if start is None:
            continue
        principal = ws[f"{principal_col}{r}"].value
        name = ws[f"{name_col}{r}"].value
        rows.append((start, principal, name))
    rows.sort(key=lambda x: x[0])
    return rows


def eval_e_value(title: object, material_map: dict[str, object]) -> str:
    """
    Reproduce:
    IF(XLOOKUP(K4, material_list_T[Title Material], material_list_T[Type Material], "-")=0, "-", XLOOKUP(...))
    """
    if title is None:
        return "-"
    result = material_map.get(str(title).strip(), "-")
    if result == 0:
        return "-"
    if result is None:
        return "-"
    return str(result)


def eval_q_value(model_writer: object, l_value: object,
                 ov_write_rows: list[tuple[date, object, object]],
                 ov_local_rows: list[tuple[date, object, object]]) -> str:
    """
    Reproduce the Q formula using Python, without modifying the workbook.
    """
    prefix = str(model_writer or "")[:2].lower()
    target = _to_date(l_value)

    if not prefix:
        return "-"

    if prefix == "ov":
        principal, name = _xlookup_last_le_pair(target, ov_write_rows) if target else ("-", "-")
    elif prefix == "lo":
        principal, name = _xlookup_last_le_pair(target, ov_local_rows) if target else ("-", "-")
    else:
        return "- - -"

    principal = "-" if principal is None else str(principal)
    name = "-" if name is None else str(name)
    return f"{principal} - {name}"


def main() -> int:
    app = QApplication.instance() or QApplication(sys.argv)

    if len(sys.argv) <= 2:
        msgbox(
            "Faltan argumentos",
            'Debes pasar el año y la versión: "1" o "2"   ',
            QMessageBox.Icon.Critical,
        )
        return 1

    year = sys.argv[1]
    version = sys.argv[2]

    if version == "1":
        filter_q_values = FILTER_Q_VALUES_1
    elif version == "2":
        filter_q_values = FILTER_Q_VALUES_2
    else:
        msgbox(
            "Argumento de versión incorrecta",
            "Señala con 1 o 2 la versión de sobrescritura.",
            QMessageBox.Icon.Critical,
        )
        return 1

    px = f"{int(year) - 2003:02d}"
    excel_path = f"E:/_Internal/{year}/{px}. identity/{px}. le_etude.overwrite.xlsx"

    # data_only=False: tomamos los datos de origen y calculamos E/Q en Python.
    wb = openpyxl.load_workbook(excel_path, data_only=False)

    if SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"La hoja '{SHEET_NAME}' no existe en el archivo.")

    ws = wb[SHEET_NAME]
    material_ws = wb["material_list"]
    ov_ws = wb["ov_models"]

    material_map = _build_material_map(material_ws)
    ov_write_rows = _build_ov_lookup(ov_ws, "B", "A", "F")
    ov_local_rows = _build_ov_lookup(ov_ws, "D", "A", "G")

    copied_values: list[str] = []

    for row in range(4, ws.max_row + 1):
        title_value = ws[f"K{row}"].value
        l_value = ws[f"L{row}"].value
        model_writer = ws[f"O{row}"].value

        e_value = eval_e_value(title_value, material_map)
        q_value = eval_q_value(model_writer, l_value, ov_write_rows, ov_local_rows)

        if e_value == FILTER_E_VALUE and q_value in filter_q_values:
            if l_value is None:
                continue
            copied_values.append(str(l_value))

    clipboard_text = "\n".join(copied_values)
    app.clipboard().setText(clipboard_text)

    msgbox(
        "Éxito",
        (
            "Los datos se obtuvieron con éxito.\n"
            f"Filtrados: {len(copied_values)}\n"
            "Copiados al portapapeles desde la columna L."
        ),
        QMessageBox.Icon.Information,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
