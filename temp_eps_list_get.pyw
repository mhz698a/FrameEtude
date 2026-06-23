import sys
from openpyxl import load_workbook
from PyQt6.QtWidgets import QApplication, QMessageBox

SHEET_NAME = "material_list"

# Cambia este valor por el filtro que necesites en la columna E.
# En el archivo actual los valores visibles son: "Episodio" y "Soundtrack".
FILTER_VALUE = "Episodio"

HEADER_ROW = 3
DATA_START_ROW = HEADER_ROW + 1

def main():
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)
    
    if not len(sys.argv) > 1:
        errmsg = QMessageBox()
        errmsg.setIcon(QMessageBox.Icon.Critical)
        errmsg.setWindowTitle("Faltan Argumentos")
        errmsg.setText(f"Añade en la ejecucion el argumento del año para poder abrirlo")
        errmsg.exec()
        exit(1)

    year = sys.argv[1]
    px = f"{int(year) - 2003:02d}"
    EXCEL_PATH = f"{'E:/_Internal'}/{year}/{px}. identity/{px}. le_etude.overwrite.xlsx"

    wb = load_workbook(EXCEL_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"No existe la hoja '{SHEET_NAME}'")

    ws = wb[SHEET_NAME]

    copied_values = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        value_e = ws[f"E{row}"].value
        if value_e == FILTER_VALUE:
            value_i = ws[f"I{row}"].value
            if value_i is not None:
                copied_values.append(str(value_i))

    text_to_copy = "\n".join(copied_values)

    clipboard = app.clipboard()
    clipboard.setText(text_to_copy)

    msg = QMessageBox()
    msg.setIcon(QMessageBox.Icon.Information)
    msg.setWindowTitle("Proceso completado")
    msg.setText(
        f"Los datos se obtuvieron con éxito.\n"
        f"Filtro aplicado en columna E: {FILTER_VALUE}\n"
        f"Registros copiados: {len(copied_values)}"
    )
    msg.exec()

if __name__ == "__main__":
    main()
